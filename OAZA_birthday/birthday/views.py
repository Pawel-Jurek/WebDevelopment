from django.http import JsonResponse
from django.shortcuts import render, redirect
from openpyxl import load_workbook
from django.views.generic import ListView

# Create your views here.
from .models import Person, Group

class PeopleListView(ListView):
    model = Person


index_view = PeopleListView.as_view()


class CurrentBirthdayListView(ListView):
    model = Person
    template_name = 'birthday/current_birthday_list.html'


def import_people_from_excel(request):
    new = 0
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file, read_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if None in row:
                break
            try:
                first_name, last_name, birthday, group_name = row
            except:
                print(row)
                continue  
            
            try:
                group = Group.objects.get(name=group_name)
            except:
                print("error")
                break
                #print(f'name: {first_name} {last_name} date: {birthday} group: {group_name}')
            
            existing_person = Person.objects.filter(
                first_name=first_name,
                last_name=last_name,
                birthday=birthday,
            ).first()

            if not existing_person:
                person = Person.objects.create(
                    first_name=first_name,
                    last_name=last_name,
                    birthday=birthday,
                    group=group,
                )
                new += 1
        print(f'New people: {new}')

        return redirect('main_page')
    
    return render(request, 'import_form.html')



def chocolate_tick(request, person_id):
    if request.method == "PUT":
        person = Person.objects.get(pk = person_id)
        person.chocolate_received = not person.chocolate_received
        person.save()
        return JsonResponse({'ticked': person.chocolate_received})
    else:
         return JsonResponse({
            "error": "PUT request required."
        }, status=400)

